import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
from database import init_db, migrate_db, load_admin_data, import_data, export_data, export_talent_pool, save_person, save_and_add_to_talent_pool, add_to_talent_pool, delete_person
from utils import check_password, save_password, validate_password, upload_photo, delete_photo, backup_data, export_person_data
import sqlite3
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side
from PIL import Image, ImageTk
import datetime
import os
import logging
import time
import re
import sys
import io

class HRManagementApp:
    def __init__(self, root):
        self.root = root
        self.root.title("人事管理系统")
        self.setup_database_and_icons()
        migrate_db()
        self.admin_data = load_admin_data()
        self.root.geometry("800x480")
        self.root.configure(bg="#F0F0F0")
        self.center_window(self.root)
        self.talent_window = None
        self.add_person_window = None
        self.edit_person_window = None
        self.export_data_window = None
        self.export_talent_window = None
        self.detail_windows = {}  # 存储人员详情页窗口的字典
        self.talent_tree = None  # 人才库的 Treeview
        self.refresh_talent_list = None  # 刷新人才库列表的方法
        self.show_password_window()

    def setup_database_and_icons(self):
        # 初始化数据库并创建icons表
        init_db()
        with sqlite3.connect('hr_data.db') as conn:
            c = conn.cursor()
            c.execute('''CREATE TABLE IF NOT EXISTS icons (
                name TEXT PRIMARY KEY,
                data BLOB
            )''')
            conn.commit()

        # 加载图标到数据库（仅在表为空时执行）
        icon_folder = r"C:\Users\HUAWEI\Desktop\222\333\icons"
        icon_files = [
            "import.png", "export.png", "add.png", "backup.png", "talent.png","password.png"
        ]
        with sqlite3.connect('hr_data.db') as conn:
            c = conn.cursor()
            c.execute("SELECT COUNT(*) FROM icons")
            if c.fetchone()[0] == 0:  # 如果表为空，加载图标
                for icon_name in icon_files:
                    icon_path = os.path.join(icon_folder, icon_name)
                    if os.path.exists(icon_path):
                        with open(icon_path, "rb") as f:
                            icon_data = f.read()
                        c.execute("INSERT OR REPLACE INTO icons (name, data) VALUES (?, ?)",
                                (icon_name, icon_data))
                        logging.info(f"图标 {icon_name} 已存入数据库")
                    else:
                        logging.warning(f"图标文件 {icon_path} 不存在")
                conn.commit()

    def load_icon_from_db(self, icon_name):
        """从数据库加载图标并返回PhotoImage对象"""
        with sqlite3.connect('hr_data.db') as conn:
            c = conn.cursor()
            c.execute("SELECT data FROM icons WHERE name=?", (icon_name,))
            result = c.fetchone()
            if result:
                icon_data = result[0]
                img = Image.open(io.BytesIO(icon_data))
                img = img.resize((20, 20), Image.Resampling.LANCZOS)
                return ImageTk.PhotoImage(img)
        logging.warning(f"数据库中未找到图标 {icon_name}")
        return None

    def center_window(self, window):
        window.update_idletasks()
        width = window.winfo_width()
        height = window.winfo_height()
        x = (window.winfo_screenwidth() // 2) - (width // 2)
        y = (window.winfo_screenheight() // 2) - (height // 2)
        window.geometry(f"{width}x{height}+{x}+{y}")

    def show_password_window(self):
        self.password_window = tk.Toplevel(self.root)
        self.password_window.title("请输入密码")
        self.password_window.geometry("300x150")
        self.password_window.configure(bg="#F0F0F0")
        self.center_window(self.password_window)
        self.password_window.transient(self.root)
        self.password_window.grab_set()

        tk.Label(self.password_window, text="密码：", font=("Roboto", 10), bg="#F0F0F0").pack(pady=10)
        self.password_entry = tk.Entry(self.password_window, show="*", width=20, font=("Roboto", 10), bd=1, relief="solid", highlightbackground="#CCCCCC", highlightthickness=1)
        self.password_entry.pack(pady=5)
        self.password_entry.bind("<Return>", lambda event: self.verify_password())

        btn = tk.Button(self.password_window, text="确认", command=self.verify_password, font=("Roboto", 10), bg="#2196F3", fg="white", bd=0, relief="flat", padx=10, pady=5)
        btn.pack(pady=10)
        btn.bind("<Enter>", lambda e: btn.config(bg="#1976D2"))
        btn.bind("<Leave>", lambda e: btn.config(bg="#2196F3"))

    def verify_password(self):
        with sqlite3.connect('hr_data.db') as conn:
            c = conn.cursor()
            c.execute("SELECT password_enabled FROM users WHERE id=1")
            result = c.fetchone()
            enabled = result[0] if result else 1

        if not enabled:
            self.password_window.destroy()
            self.create_widgets()
            return

        password = self.password_entry.get()
        if check_password(password):
            self.password_window.destroy()
            with sqlite3.connect('hr_data.db') as conn:
                c = conn.cursor()
                c.execute("SELECT password_hash FROM users WHERE id=1")
                result = c.fetchone()
                from utils import hash_password
                if result and result[0] == hash_password('123456'):
                    messagebox.showinfo("提示", "检测到默认密码，请修改密码！")
                    self.show_change_password_window()
                else:
                    self.create_widgets()
        else:
            messagebox.showerror("错误", "密码错误！")

    def show_change_password_window(self, from_main=False):
        change_window = tk.Toplevel(self.root)
        change_window.title("修改密码")
        change_window.geometry("300x300")
        change_window.configure(bg="#F0F0F0")
        self.center_window(change_window)
        change_window.transient(self.root)
        change_window.grab_set()

        with sqlite3.connect('hr_data.db') as conn:
            c = conn.cursor()
            c.execute("SELECT password_enabled FROM users WHERE id=1")
            result = c.fetchone()
            password_enabled = result[0] if result else 1

        tk.Label(change_window, text="当前密码：", font=("Roboto", 10), bg="#F0F0F0").pack(pady=5)
        current_password_entry = tk.Entry(change_window, show="*", width=20, font=("Roboto", 10), bd=1, relief="solid", highlightbackground="#CCCCCC", highlightthickness=1)
        current_password_entry.pack(pady=5)
        tk.Label(change_window, text="新密码：", font=("Roboto", 10), bg="#F0F0F0").pack(pady=5)
        new_password_entry = tk.Entry(change_window, show="*", width=20, font=("Roboto", 10), bd=1, relief="solid", highlightbackground="#CCCCCC", highlightthickness=1)
        new_password_entry.pack(pady=5)
        tk.Label(change_window, text="确认密码：", font=("Roboto", 10), bg="#F0F0F0").pack(pady=5)
        confirm_password_entry = tk.Entry(change_window, show="*", width=20, font=("Roboto", 10), bd=1, relief="solid", highlightbackground="#CCCCCC", highlightthickness=1)
        confirm_password_entry.pack(pady=5)
        strength_label = tk.Label(change_window, text="", font=("Roboto", 10), bg="#F0F0F0")
        strength_label.pack(pady=5)

        def update_strength(event=None):
            password = new_password_entry.get()
            is_valid, message = validate_password(password)
            strength_label.config(text=message)

        new_password_entry.bind("<KeyRelease>", update_strength)

        def confirm_change():
            current_password = current_password_entry.get()
            new_password = new_password_entry.get()
            confirm_password = confirm_password_entry.get()
            if not check_password(current_password):
                messagebox.showerror("错误", "当前密码错误！")
                return
            if len(new_password) < 8 or not re.search(r"[a-zA-Z]", new_password) or not re.search(r"\d", new_password):
                messagebox.showerror("错误", "新密码必须大于等于8位，且包含字母和数字！")
                return
            if new_password == confirm_password:
                save_password(new_password)
                change_window.destroy()
                messagebox.showinfo("提示", "密码修改成功，请重新登录")
                if from_main:
                    self.root.destroy()
                    os.execl(sys.executable, sys.executable, *sys.argv)
                else:
                    self.create_widgets()
            else:
                messagebox.showerror("错误", "两次输入的密码不一致！")

        def disable_password():
            if messagebox.askyesno("确认", "是否关闭密码保护？（下次登录将无需密码）"):
                with sqlite3.connect('hr_data.db') as conn:
                    c = conn.cursor()
                    c.execute("UPDATE users SET password_enabled = 0, password_hash = NULL WHERE id=1")
                    conn.commit()
                messagebox.showinfo("提示", "密码保护已关闭")
                change_window.destroy()
                if from_main:
                    self.root.destroy()
                    os.execl(sys.executable, sys.executable, *sys.argv)

        def enable_password():
            current_password = current_password_entry.get()
            new_password = new_password_entry.get()
            confirm_password = confirm_password_entry.get()
            if len(new_password) < 8 or not re.search(r"[a-zA-Z]", new_password) or not re.search(r"\d", new_password):
                messagebox.showerror("错误", "新密码必须大于等于8位，且包含字母和数字！")
                return
            if new_password == confirm_password:
                save_password(new_password)
                with sqlite3.connect('hr_data.db') as conn:
                    c = conn.cursor()
                    c.execute("UPDATE users SET password_enabled = 1 WHERE id=1")
                    conn.commit()
                messagebox.showinfo("提示", "密码保护已开启，请重新登录")
                change_window.destroy()
                if from_main:
                    self.root.destroy()
                    os.execl(sys.executable, sys.executable, *sys.argv)
            else:
                messagebox.showerror("错误", "两次输入的密码不一致！")

        button_frame = tk.Frame(change_window, bg="#F0F0F0")
        button_frame.pack(pady=10)
        confirm_btn = tk.Button(button_frame, text="完成修改", command=confirm_change, font=("Roboto", 10), bg="#2196F3", fg="white", bd=0, relief="flat", padx=10, pady=5)
        confirm_btn.pack(side=tk.LEFT, padx=5)
        confirm_btn.bind("<Enter>", lambda e: confirm_btn.config(bg="#1976D2"))
        confirm_btn.bind("<Leave>", lambda e: confirm_btn.config(bg="#2196F3"))

        if password_enabled:
            disable_btn = tk.Button(button_frame, text="关闭密码", command=disable_password, font=("Roboto", 10), bg="#FF9800", fg="white", bd=0, relief="flat", padx=10, pady=5)
            disable_btn.pack(side=tk.LEFT, padx=5)
            disable_btn.bind("<Enter>", lambda e: disable_btn.config(bg="#F57C00"))
            disable_btn.bind("<Leave>", lambda e: disable_btn.config(bg="#FF9800"))
        else:
            enable_btn = tk.Button(button_frame, text="开启密码", command=enable_password, font=("Roboto", 10), bg="#2196F3", fg="white", bd=0, relief="flat", padx=10, pady=5)
            enable_btn.pack(side=tk.LEFT, padx=5)
            enable_btn.bind("<Enter>", lambda e: enable_btn.config(bg="#1976D2"))
            enable_btn.bind("<Leave>", lambda e: enable_btn.config(bg="#2196F3"))

    def create_widgets(self):
        toolbar = tk.Frame(self.root, bg="#2196F3")
        toolbar.pack(side=tk.TOP, fill=tk.X, padx=5, pady=5)
        buttons = [
            ("导入数据", self.import_data, "import.png"),
            ("导出数据", self.export_data, "export.png"),
            ("新增人员", self.add_person, "add.png"),
            ("备份数据", self.backup_data, "backup.png"),
            ("人才库", self.show_talent_pool, "talent.png"),
        ]
        for text, command, icon in buttons:
            photo = self.load_icon_from_db(icon)
            if photo:
                btn = tk.Button(toolbar, text=text, image=photo, compound=tk.LEFT, command=command, font=("Roboto", 10), bg="#2196F3", fg="white", bd=0, relief="flat", padx=10, pady=5)
                btn.image = photo  # 保持引用
            else:
                btn = tk.Button(toolbar, text=text, command=command, font=("Roboto", 10), bg="#2196F3", fg="white", bd=0, relief="flat", padx=10, pady=5)
            btn.pack(side=tk.LEFT, padx=5)
            btn.bind("<Enter>", lambda e, b=btn: b.config(bg="#1976D2"))
            btn.bind("<Leave>", lambda e, b=btn: b.config(bg="#2196F3"))
        pwd_btn = tk.Button(toolbar, text="修改密码", command=lambda: self.show_change_password_window(from_main=True), font=("Roboto", 10), bg="#2196F3", fg="white", bd=0, relief="flat", padx=10, pady=5)
        pwd_btn.pack(side=tk.RIGHT, padx=5)
        pwd_btn.bind("<Enter>", lambda e: pwd_btn.config(bg="#1976D2"))
        pwd_btn.bind("<Leave>", lambda e: pwd_btn.config(bg="#2196F3"))

        query_frame = tk.Frame(self.root, bg="#FFFFFF", bd=1, relief="solid")
        query_frame.pack(fill=tk.X, padx=10, pady=10)
        tk.Label(query_frame, text="省份：", font=("Roboto", 10, "bold"), bg="#FFFFFF").pack(side=tk.LEFT, padx=10)
        self.province_combo = ttk.Combobox(query_frame, values=["全部"] + list(self.admin_data.keys()), width=15, font=("Roboto", 10))
        self.province_combo.set("全部")
        self.province_combo.pack(side=tk.LEFT, padx=10)
        self.province_combo.bind("<<ComboboxSelected>>", self.update_city_combo)

        tk.Label(query_frame, text="城市：", font=("Roboto", 10, "bold"), bg="#FFFFFF").pack(side=tk.LEFT, padx=10)
        self.city_combo = ttk.Combobox(query_frame, values=["全部"], width=15, font=("Roboto", 10))
        self.city_combo.set("全部")
        self.city_combo.pack(side=tk.LEFT, padx=10)

        tk.Label(query_frame, text="姓名/手机号：", font=("Roboto", 10, "bold"), bg="#FFFFFF").pack(side=tk.LEFT, padx=10)
        self.search_entry = tk.Entry(query_frame, width=15, font=("Roboto", 10), bd=1, relief="solid", highlightbackground="#CCCCCC", highlightthickness=1)
        self.search_entry.pack(side=tk.LEFT, padx=10)
        query_btn = tk.Button(query_frame, text="查询", command=self.query_by_division, font=("Roboto", 10), bg="#2196F3", fg="white", bd=0, relief="flat", padx=10, pady=5)
        query_btn.pack(side=tk.LEFT, padx=10)
        query_btn.bind("<Enter>", lambda e: query_btn.config(bg="#1976D2"))
        query_btn.bind("<Leave>", lambda e: query_btn.config(bg="#2196F3"))

        tree_frame = tk.Frame(self.root, bg="#FFFFFF")
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        style = ttk.Style()
        style.configure("Treeview", rowheight=25, font=("Roboto", 10))
        style.configure("Treeview.Heading", font=("Roboto", 10, "bold"), background="#E3F2FD")
        style.map("Treeview", background=[('selected', '#BBDEFB')], foreground=[('selected', 'black')])
        self.tree = ttk.Treeview(tree_frame, columns=("ID", "姓名", "性别", "年龄", "手机号", "省份", "城市", "分会职务", "在职状态"), show="headings", height=10)
        self.tree.heading("ID", text="ID")
        self.tree.heading("姓名", text="姓名")
        self.tree.heading("性别", text="性别")
        self.tree.heading("年龄", text="年龄")
        self.tree.heading("手机号", text="手机号")
        self.tree.heading("省份", text="省份")
        self.tree.heading("城市", text="城市")
        self.tree.heading("分会职务", text="分会职务")
        self.tree.heading("在职状态", text="在职状态")
        column_widths = {"ID": 50, "姓名": 120, "性别": 60, "年龄": 60, "手机号": 120, "省份": 100, "城市": 100, "分会职务": 120, "在职状态": 100}
        for col, width in column_widths.items():
            self.tree.column(col, width=width, anchor="center")
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.tag_configure("red", foreground="#D32F2F")
        self.tree.tag_configure("blue", foreground="#1976D2")
        self.tree.tag_configure("oddrow", background="#F5F5F5")
        self.tree.tag_configure("evenrow", background="#FFFFFF")
        self.tree.bind("<Double-1>", self.show_person_details)
        self.tree.bind("<Button-3>", self.show_popup_menu)

        self.popup_menu = tk.Menu(self.root, tearoff=0, font=("Roboto", 10))
        self.popup_menu.add_command(label="删除该人员", command=self.delete_person_from_main)

        self.refresh_data()

    def show_popup_menu(self, event):
        item = self.tree.identify_row(event.y)
        if item:
            self.tree.selection_set(item)
            self.popup_menu.post(event.x_root, event.y_root)

    def delete_person_from_main(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要删除的人员！")
            return
        item = self.tree.item(selected[0])
        person_id = item["tags"][-1]
        if messagebox.askyesno("确认", "是否彻底删除该人员？"):
            message, error = delete_person(person_id)
            if error:
                messagebox.showerror("错误", error)
            else:
                self.refresh_data()
                messagebox.showinfo("提示", message)

    def refresh_data(self):
        start_time = time.time()
        self.admin_data = load_admin_data()
        self.province_combo['values'] = ["全部"] + list(self.admin_data.keys())
        self.city_combo['values'] = ["全部"]
        for item in self.tree.get_children():
            self.tree.delete(item)
        with sqlite3.connect('hr_data.db') as conn:
            c = conn.cursor()
            c.execute("SELECT id, real_name, gender, age, phone, province, city, position, status FROM personnel")
            rows = c.fetchall()
        for idx, row in enumerate(rows, 1):
            position = row[7] if row[7] else "无职务"
            tag = "red" if row[8] == "离职" else "blue" if row[8] == "无职务" else ""
            row_tag = "oddrow" if idx % 2 else "evenrow"
            logging.info(f"人员 {row[1]}: status={row[8]}, tag={tag}")
            self.tree.insert("", "end", values=(row[0], row[1], row[2], row[3], row[4], row[5], row[6], position, row[8]), tags=(tag, row_tag, row[0]))
        elapsed_time = time.time() - start_time
        logging.info(f"主窗口数据刷新完成，记录数：{len(rows)}，耗时：{elapsed_time:.2f}秒")

    def update_city_combo(self, event=None):
        province = self.province_combo.get()
        if province == "全部":
            self.city_combo['values'] = ["全部"]
        else:
            self.city_combo['values'] = ["全部"] + self.admin_data[province]
        self.city_combo.set("全部")

    def import_data(self):
        file_paths = filedialog.askopenfilenames(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv")])
        if file_paths:
            message, error = import_data(file_paths, self.refresh_data)
            if error:
                messagebox.showerror("错误", error)
            else:
                messagebox.showinfo("导入结果", message)

    def backup_data(self):
        backup_path = filedialog.asksaveasfilename(defaultextension=".db", filetypes=[("Database files", "*.db")])
        message, error = backup_data(backup_path)
        if error:
            messagebox.showerror("错误", error)
        else:
            messagebox.showinfo("成功", message)

    def export_data(self):
        if self.export_data_window and self.export_data_window.winfo_exists():
            self.export_data_window.focus_set()
            return
        self.export_data_window = tk.Toplevel(self.root)
        self.export_data_window.title("导出数据")
        self.export_data_window.geometry("400x300")
        self.export_data_window.configure(bg="#F0F0F0")
        self.center_window(self.export_data_window)
        self.export_data_window.protocol("WM_DELETE_WINDOW", lambda: self.close_export_data_window())

        tk.Label(self.export_data_window, text="导出范围:", font=("Roboto", 10), bg="#F0F0F0").pack(pady=5)
        export_type = tk.StringVar(value="all")
        tk.Radiobutton(self.export_data_window, text="全部数据", variable=export_type, value="all", font=("Roboto", 10), bg="#F0F0F0").pack(anchor="center", padx=20)
        tk.Radiobutton(self.export_data_window, text="按分会导出", variable=export_type, value="division", font=("Roboto", 10), bg="#F0F0F0").pack(anchor="center", padx=20)

        division_frame = tk.Frame(self.export_data_window, bg="#F0F0F0")
        tk.Label(division_frame, text="省份:", font=("Roboto", 10), bg="#F0F0F0").pack(side=tk.LEFT, padx=5)
        province_combo = ttk.Combobox(division_frame, values=["全部"] + list(self.admin_data.keys()), width=15, font=("Roboto", 10))
        province_combo.set("全部")
        province_combo.pack(side=tk.LEFT, padx=5)

        tk.Label(division_frame, text="城市:", font=("Roboto", 10), bg="#F0F0F0").pack(side=tk.LEFT, padx=5)
        city_combo = ttk.Combobox(division_frame, values=["全部"], width=15, font=("Roboto", 10))
        city_combo.set("全部")
        city_combo.pack(side=tk.LEFT, padx=5)

        def update_city_combo(event=None):
            province = province_combo.get()
            if province == "全部":
                city_combo['values'] = ["全部"]
            else:
                city_combo['values'] = ["全部"] + self.admin_data[province]
            city_combo.set("全部")

        province_combo.bind("<<ComboboxSelected>>", update_city_combo)

        def update_export_options():
            if export_type.get() == "division":
                division_frame.pack(pady=5)
            else:
                division_frame.pack_forget()

        export_type.trace("w", lambda *args: update_export_options())
        update_export_options()

        def do_export():
            df, default_filename_or_error = export_data(export_type.get(), province_combo.get(), city_combo.get(), self.admin_data)
            if df is None:
                messagebox.showwarning("提示", default_filename_or_error)
                return

            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], initialfile=default_filename_or_error)
            if file_path:
                if export_type.get() == "all":
                    provinces = sorted(df['省份'].unique())
                    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                        for province in provinces:
                            province_df = df[df['省份'] == province]
                            province_df.to_excel(writer, sheet_name=f"{province}省", index=False)
                            ws = writer.sheets[f"{province}省"]
                            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                            for cell in ws[1]:
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                                cell.font = Font(name='SimSun', bold=True, size=10)
                                cell.border = thin_border
                            for col in ws.columns:
                                col_letter = col[0].column_letter
                                column_name = ws[f"{col_letter}1"].value or ''
                                title_width = max(len(str(column_name)) * 1.2, 10)
                                if column_name in ['家庭住址', '个人简历']:
                                    title_width = max(title_width, 30)
                                ws.column_dimensions[col_letter].width = title_width
                            for row in ws.iter_rows(min_row=2):
                                for cell in row:
                                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                                    cell.font = Font(name='SimSun', size=10)
                                    cell.border = thin_border
                            ws.row_dimensions[1].height = 20
                            for row in range(2, ws.max_row + 1):
                                ws.row_dimensions[row].height = 30
                else:
                    df.to_excel(file_path, index=False)
                    wb = openpyxl.load_workbook(file_path)
                    ws = wb.active
                    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    for cell in ws[1]:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font = Font(name='SimSun', bold=True, size=10)
                        cell.border = thin_border
                    for col in ws.columns:
                        col_letter = col[0].column_letter
                        column_name = ws[f"{col_letter}1"].value or ''
                        title_width = max(len(str(column_name)) * 1.2, 10)
                        if column_name in ['家庭住址', '个人简历']:
                            title_width = max(title_width, 30)
                        ws.column_dimensions[col_letter].width = title_width
                    for row in ws.iter_rows(min_row=2):
                        for cell in row:
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                            cell.font = Font(name='SimSun', size=10)
                            cell.border = thin_border
                    ws.row_dimensions[1].height = 20
                    for row in range(2, ws.max_row + 1):
                        ws.row_dimensions[row].height = 30
                    wb.save(file_path)
                messagebox.showinfo("成功", f"成功导出 {len(df)} 条数据！")
                logging.info(f"导出数据完成：{len(df)} 条")
            self.close_export_data_window()

        export_btn = tk.Button(self.export_data_window, text="确认导出", command=do_export, font=("Roboto", 10), bg="#2196F3", fg="white", bd=0, relief="flat", padx=10, pady=5)
        export_btn.pack(pady=20)
        export_btn.bind("<Enter>", lambda e: export_btn.config(bg="#1976D2"))
        export_btn.bind("<Leave>", lambda e: export_btn.config(bg="#2196F3"))

    def close_export_data_window(self):
        if self.export_data_window:
            self.export_data_window.destroy()
            self.export_data_window = None

    def query_by_division(self):
        province = self.province_combo.get().replace("省", "")
        city = self.city_combo.get().replace("市", "")
        search = self.search_entry.get()
        query = "SELECT id, real_name, gender, age, phone, province, city, position, status FROM personnel"
        params = []
        conditions = []
        if province != "全部":
            conditions.append("province LIKE ?")
            params.append(f"%{province}%")
        if city != "全部":
            conditions.append("city LIKE ?")
            params.append(f"%{city}%")
        if search:
            conditions.append("(real_name LIKE ? OR phone LIKE ?)")
            params.extend([f"%{search}%", f"%{search}%"])
        if conditions:
            query += " WHERE " + " AND ".join(conditions)

        with sqlite3.connect('hr_data.db') as conn:
            c = conn.cursor()
            c.execute(query, params)
            rows = c.fetchall()

        for item in self.tree.get_children():
            self.tree.delete(item)

        for idx, row in enumerate(rows, 1):
            position = row[7] if row[7] else "无职务"
            tag = "red" if row[8] == "离职" else "blue" if row[8] == "无职务" else ""
            row_tag = "oddrow" if idx % 2 else "evenrow"
            logging.info(f"查询人员 {row[1]}: status={row[8]}, tag={tag}")
            self.tree.insert("", "end", values=(row[0], row[1], row[2], row[3], row[4], row[5], row[6], position, row[8]), tags=(tag, row_tag, row[0]))
        logging.info(f"按分会查询完成，记录数：{len(rows)}")

    def show_talent_pool(self):
        if self.talent_window and self.talent_window.winfo_exists():
            self.talent_window.focus_set()
            return
        self.talent_window = tk.Toplevel(self.root)
        self.talent_window.title("人才库")
        self.talent_window.geometry("900x500")
        self.talent_window.configure(bg="#F0F0F0")
        self.center_window(self.talent_window)
        self.talent_window.protocol("WM_DELETE_WINDOW", lambda: self.close_talent_window())

        query_frame = tk.Frame(self.talent_window, bg="#FFFFFF", bd=1, relief="solid")
        query_frame.pack(fill=tk.X, padx=10, pady=10)
        tk.Label(query_frame, text="姓名/手机号：", font=("Roboto", 10, "bold"), bg="#FFFFFF").pack(side=tk.LEFT, padx=10)
        search_entry = tk.Entry(query_frame, width=20, font=("Roboto", 10), bd=1, relief="solid", highlightbackground="#CCCCCC", highlightthickness=1)
        search_entry.pack(side=tk.LEFT, padx=10)
        query_btn = tk.Button(query_frame, text="查询", command=lambda: self.refresh_talent_list(search_entry.get()), font=("Roboto", 10), bg="#2196F3", fg="white", bd=0, relief="flat", padx=10, pady=5)
        query_btn.pack(side=tk.LEFT, padx=10)
        query_btn.bind("<Enter>", lambda e: query_btn.config(bg="#1976D2"))
        query_btn.bind("<Leave>", lambda e: query_btn.config(bg="#2196F3"))

        tree_frame = tk.Frame(self.talent_window, bg="#FFFFFF")
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        style = ttk.Style()
        style.configure("Treeview", rowheight=25, font=("Roboto", 10))
        style.configure("Treeview.Heading", font=("Roboto", 10, "bold"), background="#E3F2FD")
        style.map("Treeview", background=[('selected', '#BBDEFB')], foreground=[('selected', 'black')])
        self.talent_tree = ttk.Treeview(tree_frame, columns=("序号", "姓名", "手机号", "省份", "城市", "分会职务", "加入人才库理由", "加入人才库时间"), show="headings", style="Treeview")
        self.talent_tree.heading("序号", text="序号")
        self.talent_tree.heading("姓名", text="姓名")
        self.talent_tree.heading("手机号", text="手机号")
        self.talent_tree.heading("省份", text="省份")
        self.talent_tree.heading("城市", text="城市")
        self.talent_tree.heading("分会职务", text="分会职务")
        self.talent_tree.heading("加入人才库理由", text="加入人才库理由")
        self.talent_tree.heading("加入人才库时间", text="加入人才库时间")
        column_widths = {"序号": 50, "姓名": 150, "手机号": 120, "省份": 100, "城市": 100, "分会职务": 120, "加入人才库理由": 100, "加入人才库时间": 150}
        for col, width in column_widths.items():
            self.talent_tree.column(col, width=width, anchor="center")
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.talent_tree.yview)
        self.talent_tree.configure(yscrollcommand=scrollbar.set)
        self.talent_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.talent_tree.bind("<Double-1>", lambda event: self.show_person_details_from_talent(event))
        self.talent_tree.tag_configure("oddrow", background="#F5F5F5")
        self.talent_tree.tag_configure("evenrow", background="#FFFFFF")

        button_frame = tk.Frame(self.talent_window, bg="#F0F0F0")
        button_frame.pack(pady=15)
        remove_btn = tk.Button(button_frame, text="从人才库中移除", command=self.remove_selected, font=("Roboto", 10), bg="#FF9800", fg="white", bd=0, relief="flat", padx=10, pady=5)
        remove_btn.pack(side=tk.LEFT, padx=10)
        remove_btn.bind("<Enter>", lambda e: remove_btn.config(bg="#F57C00"))
        remove_btn.bind("<Leave>", lambda e: remove_btn.config(bg="#FF9800"))
        export_btn = tk.Button(button_frame, text="导出人才库", command=self.export_talent_pool_with_confirm, font=("Roboto", 10), bg="#2196F3", fg="white", bd=0, relief="flat", padx=10, pady=5)
        export_btn.pack(side=tk.LEFT, padx=10)
        export_btn.bind("<Enter>", lambda e: export_btn.config(bg="#1976D2"))
        export_btn.bind("<Leave>", lambda e: export_btn.config(bg="#2196F3"))

        def refresh_talent_list(search=""):
            for item in self.talent_tree.get_children():
                self.talent_tree.delete(item)
            with sqlite3.connect('hr_data.db') as conn:
                c = conn.cursor()
                query = """
                    SELECT p.id, p.real_name, p.phone, p.province, p.city, p.position, t.reason, t.add_time
                    FROM personnel p
                    JOIN talent_pool t ON p.id = t.person_id
                """
                params = []
                if search:
                    query += " WHERE p.real_name LIKE ? OR p.phone LIKE ?"
                    params = [f"%{search}%", f"%{search}%"]
                query += " ORDER BY t.add_time DESC"
                c.execute(query, params)
                rows = c.fetchall()
            logging.info(f"人才库查询结果：{len(rows)} 条记录")
            for idx, row in enumerate(rows, 1):
                reason = row[6] if row[6] else "无"
                position = row[5] if row[5] else "无职务"
                row_tag = "oddrow" if idx % 2 else "evenrow"
                self.talent_tree.insert("", "end", values=(idx, row[1], row[2], row[3], row[4], position, reason, row[7]), tags=(row_tag, row[0]))

        self.refresh_talent_list = refresh_talent_list
        self.refresh_talent_list()

    def remove_selected(self):
        selected = self.talent_tree.selection()
        if not selected:
            messagebox.showwarning("提示", "请先选择要移除的人员！")
            return
        if messagebox.askyesno("确认", "是否从人才库中移除选中人员？（数据库中保留）"):
            with sqlite3.connect('hr_data.db') as conn:
                c = conn.cursor()
                for item in selected:
                    person_id = self.talent_tree.item(item, "tags")[1]
                    c.execute("SELECT real_name FROM personnel WHERE id=?", (person_id,))
                    real_name = c.fetchone()[0]
                    c.execute("DELETE FROM talent_pool WHERE person_id=?", (person_id,))
                    c.execute("INSERT INTO operation_log (operation_type, operation_target, operation_time) VALUES (?, ?, ?)",
                              ("从人才库中移除", real_name, datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
                conn.commit()
            self.refresh_talent_list()
            messagebox.showinfo("提示", "已从人才库中移除选中人员")
            logging.info("从人才库中移除人员完成")

    def close_talent_window(self):
        if self.talent_window:
            self.talent_window.destroy()
            self.talent_window = None
            self.talent_tree = None
            self.refresh_talent_list = None

    def export_talent_pool_with_confirm(self):
        if self.export_talent_window and self.export_talent_window.winfo_exists():
            self.export_talent_window.focus_set()
            return
        self.export_talent_window = tk.Toplevel(self.root)
        self.export_talent_window.title("导出人才库")
        self.export_talent_window.geometry("300x150")
        self.export_talent_window.configure(bg="#F0F0F0")
        self.center_window(self.export_talent_window)
        self.export_talent_window.protocol("WM_DELETE_WINDOW", lambda: self.close_export_talent_window())

        tk.Label(self.export_talent_window, text="是否导出人才库全部人员？", font=("Roboto", 10), bg="#F0F0F0").pack(pady=20)
        button_frame = tk.Frame(self.export_talent_window, bg="#F0F0F0")
        button_frame.pack(pady=10)
        confirm_btn = tk.Button(button_frame, text="确认", command=self.do_export_talent_pool, font=("Roboto", 10), bg="#2196F3", fg="white", bd=0, relief="flat", padx=10, pady=5)
        confirm_btn.pack(side=tk.LEFT, padx=5)
        confirm_btn.bind("<Enter>", lambda e: confirm_btn.config(bg="#1976D2"))
        confirm_btn.bind("<Leave>", lambda e: confirm_btn.config(bg="#2196F3"))
        cancel_btn = tk.Button(button_frame, text="取消", command=self.close_export_talent_window, font=("Roboto", 10), bg="#FF9800", fg="white", bd=0, relief="flat", padx=10, pady=5)
        cancel_btn.pack(side=tk.LEFT, padx=5)
        cancel_btn.bind("<Enter>", lambda e: cancel_btn.config(bg="#F57C00"))
        cancel_btn.bind("<Leave>", lambda e: cancel_btn.config(bg="#FF9800"))

    def do_export_talent_pool(self):
        df, error = export_talent_pool()
        if error:
            messagebox.showerror("错误", error)
            return
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], initialfile="人才库名单")
        if file_path:
            df.to_excel(file_path, index=False)
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for cell in ws[1]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(name='SimSun', bold=True, size=10)
                cell.border = thin_border
            for col in ws.columns:
                col_letter = col[0].column_letter
                column_name = ws[f"{col_letter}1"].value or ''
                title_width = max(len(str(column_name)) * 1.2, 10)
                if column_name in ['个人简历', '加入人才库理由']:
                    title_width = max(title_width, 30)
                ws.column_dimensions[col_letter].width = title_width
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    cell.font = Font(name='SimSun', size=10)
                    cell.border = thin_border
            ws.row_dimensions[1].height = 20
            for row in range(2, ws.max_row + 1):
                ws.row_dimensions[row].height = 30
            wb.save(file_path)
            messagebox.showinfo("成功", f"成功导出人才库名单！")
            logging.info("导出人才库完成")
        self.close_export_talent_window()

    def close_export_talent_window(self):
        if self.export_talent_window:
            self.export_talent_window.destroy()
            self.export_talent_window = None

    def show_person_details_from_talent(self, event):
        selected = self.talent_tree.selection()
        if not selected:
            return
        item = self.talent_tree.item(selected[0])
        person_id = item["tags"][-1]
        logging.info(f"人才库人员详情：tags={item['tags']}, person_id={person_id}")
        with sqlite3.connect('hr_data.db') as conn:
            c = conn.cursor()
            c.execute("SELECT * FROM personnel WHERE id=?", (person_id,))
            person = c.fetchone()
        if person:
            self.show_person_details_manual(person, from_talent=True)
        else:
            messagebox.showerror("错误", "未找到该人员信息！")
            logging.error(f"人才库人员详情查询失败：ID {person_id}")

    def show_person_details(self, event):
        selected = self.tree.selection()
        if not selected:
            return
        item = self.tree.item(selected[0])
        person_id = item["tags"][-1]
        logging.info(f"主窗口人员详情：tags={item['tags']}, person_id={person_id}")
        with sqlite3.connect('hr_data.db') as conn:
            c = conn.cursor()
            c.execute("SELECT * FROM personnel WHERE id=?", (person_id,))
            person = c.fetchone()
        if person:
            self.show_person_details_manual(person)
        else:
            messagebox.showerror("错误", "未找到该人员信息！")
            logging.error(f"人员详情查询失败：ID {person_id}")

    def show_person_details_manual(self, person, from_talent=False):
        person_id = person[0]
        key = (person_id, from_talent)
        if key in self.detail_windows and self.detail_windows[key].winfo_exists():
            self.detail_windows[key].focus_set()
            return

        detail_window = tk.Toplevel(self.root)
        detail_window.title("人员详情")
        detail_window.geometry("505x640")
        detail_window.configure(bg="#F0F0F0")
        self.center_window(detail_window)

        self.detail_windows[key] = detail_window

        def on_close():
            if key in self.detail_windows:
                del self.detail_windows[key]
            detail_window.destroy()

        detail_window.protocol("WM_DELETE_WINDOW", on_close)

        main_frame = tk.Frame(detail_window, bg="#FFFFFF", bd=2, relief="solid")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        basic_frame = tk.LabelFrame(main_frame, text="基本信息", font=("Roboto", 10, "bold"), bg="#FFFFFF")
        basic_frame.pack(fill=tk.X, padx=5, pady=5)
        basic_inner = tk.Frame(basic_frame, bg="#FFFFFF")
        basic_inner.pack(fill=tk.X, padx=5, pady=5)

        basic_inner.grid_columnconfigure(0, minsize=100)
        basic_inner.grid_columnconfigure(1, minsize=180)
        basic_inner.grid_columnconfigure(2, weight=0, minsize=150)

        basic_fields = [("真实姓名", 1), ("性别", 2), ("年龄", 3), ("身份证号", 4), ("手机号", 5), ("分会职务", 13), ("在职状态", 14)]
        for i, (field, idx) in enumerate(basic_fields):
            tk.Label(basic_inner, text=f"{field}：", width=14, anchor="e", font=("Roboto", 10), bg="#FFFFFF").grid(row=i, column=0, sticky="e")
            value = person[idx] if person[idx] else "无"
            wraplength = 180 if field == "分会职务" else 0
            tk.Label(basic_inner, text=value, anchor="w", wraplength=wraplength, font=("Roboto", 10), bg="#FFFFFF").grid(row=i, column=1, sticky="w")
            if wraplength and len(value) > 12:
                logging.info(f"分会职务换行：{value}, 长度={len(value)}")

        photo_wrapper = tk.Frame(basic_inner, bg="#FFFFFF")
        photo_wrapper.grid(row=0, column=2, rowspan=7, sticky="e", padx=(10, 10), pady=5)
        photo_frame = tk.Frame(photo_wrapper, width=100, height=130, bd=1, relief="solid", bg="#F5F5F5", highlightbackground="#CCCCCC", highlightthickness=1)
        photo_frame.pack(expand=True, fill="y")
        photo_frame.pack_propagate(False)
        photo_label = tk.Label(photo_frame, bg="#F5F5F5")
        photo_label.pack(fill="both")
        has_photo = person[-1] and os.path.exists(person[-1])
        if has_photo:
            try:
                img = Image.open(person[-1])
                img.thumbnail((100, 130), Image.Resampling.LANCZOS)
                photo = ImageTk.PhotoImage(img)
                photo_label.config(image=photo)
                photo_label.image = photo
                photo_label.update_idletasks()
                logging.info(f"详情页加载照片：{person[-1]}")
            except Exception as e:
                messagebox.showwarning("警告", f"无法加载照片：{str(e)}")
                logging.error(f"加载照片失败：{str(e)}")
        else:
            tk.Label(photo_frame, text="无照片", font=("Roboto", 10), bg="#F5F5F5").pack(expand=True)

        detail_frame = tk.LabelFrame(main_frame, text="详细信息", font=("Roboto", 10, "bold"), bg="#FFFFFF")
        detail_frame.pack(fill=tk.X, padx=5, pady=5)
        detail_inner = tk.Frame(detail_frame, bg="#FFFFFF")
        detail_inner.pack(fill=tk.X, padx=10, pady=10)

        detail_left = tk.Frame(detail_inner, bg="#FFFFFF")
        detail_left.pack(side=tk.LEFT, fill=tk.Y)
        detail_left.grid_columnconfigure(0, minsize=100)
        detail_left.grid_columnconfigure(1, minsize=180)
        detail_fields = [("省份", 6), ("城市", 7), ("昵称", 9), ("学历", 10), ("政治面貌", 11),
                        ("个人职业", 12), ("加入组织时间", 15), ("跟捐天数", 16), ("家庭住址", 17)]
        for i, (field, idx) in enumerate(detail_fields):
            tk.Label(detail_left, text=f"{field}：", width=14, anchor="e", font=("Roboto", 10), bg="#FFFFFF").grid(row=i, column=0, sticky="e")
            value = person[idx] if person[idx] else "无"
            wraplength = 180 if field in ["个人职业", "家庭住址"] else 250
            tk.Label(detail_left, text=value, anchor="w", wraplength=wraplength, font=("Roboto", 10), bg="#FFFFFF").grid(row=i, column=1, sticky="w")
            if wraplength == 180 and len(value) > 12:
                logging.info(f"{field}换行：{value}, 长度={len(value)}")

        if from_talent:
            detail_right = tk.Frame(detail_inner, bg="#FFFFFF")
            detail_right.pack(side=tk.RIGHT, fill=tk.Y, padx=(5, 10))
            reason_title = tk.Label(detail_right, text="加入人才库理由", font=("Roboto", 10, "bold"), bg="#FFFFFF", anchor="center")
            reason_title.pack(fill=tk.X, pady=(0, 5))
            with sqlite3.connect('hr_data.db') as conn:
                c = conn.cursor()
                c.execute("SELECT reason FROM talent_pool WHERE person_id=?", (person[0],))
                reason = c.fetchone()
            reason_frame = tk.Frame(detail_right, width=150, height=130, bg="#F5F5F5")
            reason_frame.pack(fill=tk.X, padx=5, pady=5)
            reason_frame.pack_propagate(False)

            reason_text = tk.Text(
                reason_frame, width=18, height=7, wrap=tk.WORD, font=("Roboto", 10),
                bg="#F5F5F5", bd=0, highlightthickness=0, relief="flat"
            )
            reason_text.pack(side=tk.LEFT, padx=5, pady=5, fill="both", expand=True)
            scrollbar = ttk.Scrollbar(reason_frame, orient=tk.VERTICAL, command=reason_text.yview)
            reason_text.configure(yscrollcommand=scrollbar.set)
            reason_text.insert(tk.END, reason[0] if reason and reason[0] else "无")
            reason_text.config(state="disabled")
            line_count = reason_text.count("1.0", tk.END, "displaylines")[0]
            if line_count > 7:
                scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            if reason and reason[0] and len(reason[0]) > 12:
                logging.info(f"加入人才库理由换行：{reason[0]}, 长度={len(reason[0])}")

        bio_frame = tk.LabelFrame(main_frame, text="个人简历", font=("Roboto", 10, "bold"), bg="#FFFFFF")
        bio_frame.pack(fill=tk.BOTH, padx=5, pady=5)
        bio_inner = tk.Frame(bio_frame, bg="#F5F5F5")
        bio_inner.pack(fill=tk.BOTH, padx=5, pady=5, expand=True)

        bio_text = tk.Text(
            bio_inner, width=45, height=5, wrap=tk.WORD, font=("Roboto", 10),
            bg="#F5F5F5", bd=0, highlightthickness=0, relief="flat"
        )
        bio_content = person[-2] if person[-2] else "无"
        logging.info(f"个人简历内容：{bio_content}")
        bio_text.insert(tk.END, bio_content)
        bio_text.config(state="disabled")
        bio_text.pack(side=tk.LEFT, padx=5, pady=5, fill="both", expand=True)
        bio_text.update_idletasks()
        line_count = bio_text.count("1.0", tk.END, "displaylines")[0]
        logging.info(f"个人简历行数：{line_count}")
        if line_count >= 5:
            scrollbar = ttk.Scrollbar(bio_inner, orient=tk.VERTICAL, command=bio_text.yview)
            bio_text.configure(yscrollcommand=scrollbar.set)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        button_frame = tk.Frame(main_frame, bg="#FFFFFF")
        button_frame.pack(fill=tk.X, pady=10)
        inner_button_frame = tk.Frame(button_frame, bg="#FFFFFF")
        inner_button_frame.pack(anchor="center")
        edit_btn = tk.Button(inner_button_frame, text="修改信息", command=lambda: self.edit_person(person[0], detail_window, from_talent), font=("Roboto", 10), bg="#2196F3", fg="white", bd=0, relief="flat", padx=10, pady=5)
        edit_btn.pack(side=tk.LEFT, padx=5)
        edit_btn.bind("<Enter>", lambda e: edit_btn.config(bg="#1976D2"))
        edit_btn.bind("<Leave>", lambda e: edit_btn.config(bg="#2196F3"))
        export_btn = tk.Button(inner_button_frame, text="导出信息", command=lambda: self.export_person_data(person, from_talent), font=("Roboto", 10), bg="#2196F3", fg="white", bd=0, relief="flat", padx=10, pady=5)
        export_btn.pack(side=tk.LEFT, padx=5)
        export_btn.bind("<Enter>", lambda e: export_btn.config(bg="#1976D2"))
        export_btn.bind("<Leave>", lambda e: export_btn.config(bg="#2196F3"))
        if not from_talent:
            delete_btn = tk.Button(inner_button_frame, text="删除人员", command=lambda: self.delete_person(person[0], detail_window), font=("Roboto", 10), bg="#FF9800", fg="white", bd=0, relief="flat", padx=10, pady=5)
            delete_btn.pack(side=tk.LEFT, padx=5)
            delete_btn.bind("<Enter>", lambda e: delete_btn.config(bg="#F57C00"))
            delete_btn.bind("<Leave>", lambda e: delete_btn.config(bg="#FF9800"))
            with sqlite3.connect('hr_data.db') as conn:
                c = conn.cursor()
                c.execute("SELECT id FROM talent_pool WHERE person_id=?", (person[0],))
                is_in_talent = c.fetchone() is not None
            add_btn = tk.Button(inner_button_frame, text="加入人才库", command=lambda: self.show_reason_window(person[0], detail_window), font=("Roboto", 10), bg="#2196F3", fg="white", bd=0, relief="flat", padx=10, pady=5, state="disabled" if is_in_talent else "normal")
            add_btn.pack(side=tk.LEFT, padx=5)
            add_btn.bind("<Enter>", lambda e: add_btn.config(bg="#1976D2"))
            add_btn.bind("<Leave>", lambda e: add_btn.config(bg="#2196F3"))

    def show_reason_window(self, person_id, detail_window=None):
        reason_window = tk.Toplevel(self.root)
        reason_window.title("加入理由")
        reason_window.geometry("400x300")
        reason_window.configure(bg="#F0F0F0")
        self.center_window(reason_window)
        reason_window.transient(self.root)
        reason_window.grab_set()

        tk.Label(reason_window, text="请输入加入理由：", font=("Roboto", 10), bg="#F0F0F0").pack(pady=5)
        reason_text = scrolledtext.ScrolledText(reason_window, height=10, width=40, wrap=tk.WORD, font=("Roboto", 10), bd=1, relief="solid", highlightbackground="#CCCCCC", highlightthickness=1)
        reason_text.pack(pady=5)

        button_frame = tk.Frame(reason_window, bg="#F0F0F0")
        button_frame.pack(pady=10)
        confirm_btn = tk.Button(button_frame, text="确认加入", command=lambda: self.add_to_talent_pool(person_id, reason_text.get("1.0", tk.END).strip(), detail_window, reason_window), font=("Roboto", 10), bg="#2196F3", fg="white", bd=0, relief="flat", padx=10, pady=5)
        confirm_btn.pack(side=tk.LEFT, padx=5)
        confirm_btn.bind("<Enter>", lambda e: confirm_btn.config(bg="#1976D2"))
        confirm_btn.bind("<Leave>", lambda e: confirm_btn.config(bg="#2196F3"))
        cancel_btn = tk.Button(button_frame, text="取消", command=reason_window.destroy, font=("Roboto", 10), bg="#FF9800", fg="white", bd=0, relief="flat", padx=10, pady=5)
        cancel_btn.pack(side=tk.LEFT, padx=5)
        cancel_btn.bind("<Enter>", lambda e: cancel_btn.config(bg="#F57C00"))
        cancel_btn.bind("<Leave>", lambda e: cancel_btn.config(bg="#FF9800"))

    def add_to_talent_pool(self, person_id, reason, detail_window, reason_window):
        message, error = add_to_talent_pool(person_id, reason)
        if error:
            messagebox.showerror("错误", error)
        else:
            self.refresh_data()
            if self.talent_window and self.talent_window.winfo_exists():
                self.refresh_talent_list()
            messagebox.showinfo("提示", message)
            reason_window.destroy()
            if detail_window:
                detail_window.destroy()

    def delete_person(self, person_id, detail_window=None):
        if messagebox.askyesno("确认", "是否彻底删除该人员？"):
            message, error = delete_person(person_id)
            if error:
                messagebox.showerror("错误", error)
            else:
                self.refresh_data()
                messagebox.showinfo("提示", message)
                if detail_window:
                    detail_window.destroy()

    def edit_person(self, person_id, detail_window=None, from_talent=False):
        if self.edit_person_window and self.edit_person_window.winfo_exists():
            self.edit_person_window.focus_set()
            return
        if detail_window:
            detail_window.destroy()
        with sqlite3.connect('hr_data.db') as conn:
            c = conn.cursor()
            c.execute("SELECT * FROM personnel WHERE id=?", (person_id,))
            person = c.fetchone()
            if from_talent:
                c.execute("SELECT reason FROM talent_pool WHERE person_id=?", (person_id,))
                reason = c.fetchone()
                reason = reason[0] if reason else ""
            else:
                reason = ""
        self.open_person_window(mode="edit", person=person, from_talent=from_talent, talent_reason=reason, detail_person=person if detail_window else None)

    def add_person(self):
        if self.add_person_window and self.add_person_window.winfo_exists():
            self.add_person_window.focus_set()
            return
        self.open_person_window(mode="add")

    def upload_photo(self):
        current_focus = self.root.focus_get()
        file_path = filedialog.askopenfilename(filetypes=[("Image files", "*.png *.jpg *.jpeg")])
        if file_path:
            new_path = upload_photo(file_path)
            if new_path:
                self.photo_path.set(new_path)
                try:
                    img = Image.open(new_path)
                    img.thumbnail((100, 130), Image.Resampling.LANCZOS)
                    photo = ImageTk.PhotoImage(img)
                    self.photo_label.config(image=photo)
                    self.photo_label.image = photo
                    self.photo_label.update_idletasks()
                    logging.info(f"照片上传成功：{new_path}")
                except Exception as e:
                    messagebox.showwarning("警告", f"无法加载照片：{str(e)}")
                    logging.error(f"照片加载失败：{str(e)}")
        if current_focus:
            current_focus.focus_set()

    def open_person_window(self, mode, person=None, from_talent=False, talent_reason="", detail_person=None):
        if mode == "add":
            if self.add_person_window and self.add_person_window.winfo_exists():
                self.add_person_window.focus_set()
                return
        elif mode == "edit":
            if self.edit_person_window and self.edit_person_window.winfo_exists():
                self.edit_person_window.focus_set()
                return
        window_title = "新增人员" if mode == "add" else "编辑人员"
        window = tk.Toplevel(self.root)
        window.title(window_title)
        window.geometry("505x640")
        window.configure(bg="#F0F0F0")
        self.center_window(window)
        if mode == "add":
            self.add_person_window = window
            window.protocol("WM_DELETE_WINDOW", lambda: self.close_add_person_window())
        else:
            self.edit_person_window = window
            window.protocol("WM_DELETE_WINDOW", lambda: self.close_edit_person_window())

        main_frame = tk.Frame(window, bg="#FFFFFF", bd=2, relief="solid")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        basic_frame = tk.LabelFrame(main_frame, text="基本信息", font=("Roboto", 10, "bold"), bg="#FFFFFF")
        basic_frame.pack(fill=tk.X, padx=5, pady=5)
        basic_inner = tk.Frame(basic_frame, bg="#FFFFFF")
        basic_inner.pack(fill=tk.X, padx=5, pady=5)

        basic_inner.grid_columnconfigure(0, minsize=100)
        basic_inner.grid_columnconfigure(1, minsize=180)
        basic_inner.grid_columnconfigure(2, weight=0, minsize=150)

        basic_fields = [
            ("真实姓名", tk.Entry, None, 20),
            ("性别", ttk.Combobox, ["男", "女"], 10),
            ("年龄", tk.Entry, None, 10),
            ("身份证号", tk.Entry, None, 20),
            ("手机号", tk.Entry, None, 20),
            ("分会职务", tk.Entry, None, 20),
            ("在职状态", ttk.Combobox, ["在职", "离职", "无职务"], 10),
        ]
        entries = {}
        for i, (field, widget_type, values, width) in enumerate(basic_fields):
            tk.Label(basic_inner, text=f"{field}：", width=14, anchor="e", font=("Roboto", 10), bg="#FFFFFF").grid(row=i, column=0, sticky="e")
            entry = widget_type(basic_inner, width=width, font=("Roboto", 10))
            if values:
                entry['values'] = values
            entry.grid(row=i, column=1, sticky="w")
            if widget_type == tk.Entry:
                entry.config(bd=1, relief="solid", highlightbackground="#CCCCCC", highlightthickness=1)
            entries[field] = entry

        photo_wrapper = tk.Frame(basic_inner, bg="#FFFFFF")
        photo_wrapper.grid(row=0, column=2, rowspan=7, sticky="e", padx=(10, 10), pady=5)
        photo_frame = tk.Frame(photo_wrapper, width=100, height=130, bd=1, relief="solid", bg="#F5F5F5", highlightbackground="#CCCCCC", highlightthickness=1)
        photo_frame.pack(expand=True, fill="y")
        photo_frame.pack_propagate(False)
        self.photo_label = tk.Label(photo_frame, bg="#F5F5F5")
        self.photo_label.pack(fill="both")
        self.photo_path = tk.StringVar()
        if person and person[-1] and os.path.exists(person[-1]):
            self.photo_path.set(person[-1])
            try:
                img = Image.open(person[-1])
                img.thumbnail((100, 130), Image.Resampling.LANCZOS)
                photo = ImageTk.PhotoImage(img)
                self.photo_label.config(image=photo)
                self.photo_label.image = photo
                self.photo_label.update_idletasks()
                logging.info(f"加载现有照片：{person[-1]}")
            except Exception as e:
                messagebox.showwarning("警告", f"无法加载照片：{str(e)}")
                logging.error(f"照片加载失败：{str(e)}")
        else:
            tk.Label(self.photo_label, text="无照片", font=("Roboto", 10), bg="#F5F5F5").pack(expand=True)

        button_frame = tk.Frame(photo_wrapper, bg="#FFFFFF")
        button_frame.pack(fill=tk.X, pady=5)
        upload_btn = tk.Button(button_frame, text="上传照片", command=self.upload_photo, font=("Roboto", 10), bg="#2196F3", fg="white", bd=0, relief="flat", padx=10, pady=5)
        upload_btn.pack(side=tk.LEFT, padx=5)
        upload_btn.bind("<Enter>", lambda e: upload_btn.config(bg="#1976D2"))
        upload_btn.bind("<Leave>", lambda e: upload_btn.config(bg="#2196F3"))
        if mode == "edit":
            delete_btn = tk.Button(button_frame, text="删除照片", command=lambda: delete_photo(self.photo_path, self.photo_label), font=("Roboto", 10), bg="#FF9800", fg="white", bd=0, relief="flat", padx=10, pady=5)
            delete_btn.pack(side=tk.LEFT, padx=5)
            delete_btn.bind("<Enter>", lambda e: delete_btn.config(bg="#F57C00"))
            delete_btn.bind("<Leave>", lambda e: delete_btn.config(bg="#FF9800"))

        detail_frame = tk.LabelFrame(main_frame, text="详细信息", font=("Roboto", 10, "bold"), bg="#FFFFFF")
        detail_frame.pack(fill=tk.X, padx=5, pady=5)
        detail_inner = tk.Frame(detail_frame, bg="#FFFFFF")
        detail_inner.pack(fill=tk.X, padx=10, pady=10)

        detail_left = tk.Frame(detail_inner, bg="#FFFFFF")
        detail_left.pack(side=tk.LEFT, fill=tk.Y)
        detail_fields = [
            ("省份", tk.Entry, None, 20),
            ("城市", tk.Entry, None, 20),
            ("昵称", tk.Entry, None, 20),
            ("学历", ttk.Combobox, ["初中", "高中", "大专", "本科", "硕士", "博士"], 10),
            ("政治面貌", ttk.Combobox, ["中共党员", "共青团员", "群众"], 10),
            ("个人职业", tk.Entry, None, 20),
            ("加入组织时间", tk.Entry, None, 20),
            ("跟捐天数", tk.Entry, None, 20),
            ("家庭住址", tk.Entry, None, 20),
        ]
        for i, (field, widget_type, values, width) in enumerate(detail_fields):
            tk.Label(detail_left, text=f"{field}：", width=14, anchor="e", font=("Roboto", 10), bg="#FFFFFF").grid(row=i, column=0, sticky="e")
            entry = widget_type(detail_left, width=width, font=("Roboto", 10))
            if values:
                entry['values'] = values
            entry.grid(row=i, column=1, sticky="w")
            if widget_type == tk.Entry:
                entry.config(bd=1, relief="solid", highlightbackground="#CCCCCC", highlightthickness=1)
            entries[field] = entry

        if from_talent:
            detail_right = tk.Frame(detail_inner, bg="#FFFFFF")
            detail_right.pack(side=tk.RIGHT, fill=tk.Y, padx=(5, 10))
            tk.Label(detail_right, text="加入人才库理由", font=("Roboto", 10, "bold"), bg="#FFFFFF", anchor="center").pack(fill=tk.X, pady=(0, 5))
            reason_frame = tk.Frame(detail_right, bg="#F5F5F5", highlightbackground="#CCCCCC", highlightthickness=1)
            reason_frame.pack(fill=tk.X)
            reason_text = scrolledtext.ScrolledText(reason_frame, height=7, width=15, wrap=tk.WORD, font=("Roboto", 10), bd=1, relief="solid", highlightbackground="#CCCCCC", highlightthickness=1)
            reason_text.pack(padx=15, pady=5)
            reason_text.insert(tk.END, talent_reason if talent_reason else "")
            entries["加入人才库理由"] = reason_text

        bio_frame = tk.LabelFrame(main_frame, text="个人简历", font=("Roboto", 10, "bold"), bg="#FFFFFF")
        bio_frame.pack(fill=tk.BOTH, padx=5, pady=5)
        bio_text = scrolledtext.ScrolledText(bio_frame, height=5, width=45, wrap=tk.WORD, font=("Roboto", 10), bd=1, relief="solid", highlightbackground="#CCCCCC", highlightthickness=1)
        bio_text.pack(fill=tk.X, padx=5, pady=5)
        entries["个人简历"] = bio_text

        if mode == "edit" and person:
            entries["真实姓名"].insert(0, person[1] if person[1] else "")
            entries["性别"].set(person[2] if person[2] else "")
            entries["年龄"].insert(0, person[3] if person[3] else "")
            entries["身份证号"].insert(0, person[4] if person[4] else "")
            entries["手机号"].insert(0, person[5] if person[5] else "")
            entries["省份"].insert(0, person[6] if person[6] else "")
            entries["城市"].insert(0, person[7] if person[7] else "")
            entries["昵称"].insert(0, person[9] if person[9] else "")
            entries["学历"].set(person[10] if person[10] else "")
            entries["政治面貌"].set(person[11] if person[11] else "")
            entries["个人职业"].insert(0, person[12] if person[12] else "")
            entries["分会职务"].insert(0, person[13] if person[13] else "")
            entries["在职状态"].set(person[14] if person[14] else "")
            entries["加入组织时间"].insert(0, person[15] if person[15] else "")
            entries["跟捐天数"].insert(0, person[16] if person[16] else "")
            entries["家庭住址"].insert(0, person[17] if person[17] else "")
            bio_text.insert(tk.END, person[18] if person[18] else "")

        def save_data():
            data = [
                entries["真实姓名"].get() or None,
                entries["性别"].get() or None,
                entries["年龄"].get() or None,
                entries["身份证号"].get() or None,
                entries["手机号"].get() or None,
                entries["省份"].get() or None,
                entries["城市"].get() or None,
                None,
                entries["昵称"].get() or None,
                entries["学历"].get() or None,
                entries["政治面貌"].get() or None,
                entries["个人职业"].get() or None,
                entries["分会职务"].get() or None,
                entries["在职状态"].get() or None,
                entries["加入组织时间"].get() or None,
                entries["跟捐天数"].get() or None,
                entries["家庭住址"].get() or None,
                entries["个人简历"].get("1.0", tk.END).strip() or None,
                self.photo_path.get() or None,
            ]
            if not data[0] or not data[4]:
                messagebox.showerror("错误", "真实姓名和手机号为必填项！")
                return
            try:
                if mode == "add":
                    if from_talent:
                        reason = entries["加入人才库理由"].get("1.0", tk.END).strip() or None
                        person_id, message, error = save_and_add_to_talent_pool(data, reason)
                    else:
                        person_id, message, error = save_person(data, mode, None)
                else:
                    person_id, message, error = save_person(data, mode, person, from_talent)
                    if from_talent and "加入人才库理由" in entries:
                        reason = entries["加入人才库理由"].get("1.0", tk.END).strip() or None
                        if reason:
                            with sqlite3.connect('hr_data.db') as conn:
                                c = conn.cursor()
                                c.execute("UPDATE talent_pool SET reason = ? WHERE person_id = ?", (reason, person_id))
                                conn.commit()
                if error:
                    messagebox.showerror("错误", error)
                else:
                    self.refresh_data()
                    messagebox.showinfo("提示", message)
                    window.destroy()
                    if person_id:
                        with sqlite3.connect('hr_data.db') as conn:
                            c = conn.cursor()
                            c.execute("SELECT * FROM personnel WHERE id=?", (person_id,))
                            updated_person = c.fetchone()
                        if updated_person:
                            self.show_person_details_manual(updated_person, from_talent)
            except Exception as e:
                messagebox.showerror("错误", f"保存失败：{str(e)}")

        button_frame = tk.Frame(main_frame, bg="#FFFFFF")
        button_frame.pack(fill=tk.X, pady=10)
        inner_button_frame = tk.Frame(button_frame, bg="#FFFFFF")
        inner_button_frame.pack(anchor="center")
        save_btn = tk.Button(inner_button_frame, text="保存", command=save_data, font=("Roboto", 10), bg="#2196F3", fg="white", bd=0, relief="flat", padx=10, pady=5)
        save_btn.pack(side=tk.LEFT, padx=5)
        save_btn.bind("<Enter>", lambda e: save_btn.config(bg="#1976D2"))
        save_btn.bind("<Leave>", lambda e: save_btn.config(bg="#2196F3"))
        cancel_btn = tk.Button(inner_button_frame, text="取消", command=window.destroy, font=("Roboto", 10), bg="#FF9800", fg="white", bd=0, relief="flat", padx=10, pady=5)
        cancel_btn.pack(side=tk.LEFT, padx=5)
        cancel_btn.bind("<Enter>", lambda e: cancel_btn.config(bg="#F57C00"))
        cancel_btn.bind("<Leave>", lambda e: cancel_btn.config(bg="#FF9800"))
        if mode == "add" and not from_talent:
            save_add_btn = tk.Button(inner_button_frame, text="保存并加入人才库", command=lambda: self.save_and_add_to_talent_pool(entries, window), font=("Roboto", 10), bg="#2196F3", fg="white", bd=0, relief="flat", padx=10, pady=5)
            save_add_btn.pack(side=tk.LEFT, padx=5)
            save_add_btn.bind("<Enter>", lambda e: save_add_btn.config(bg="#1976D2"))
            save_add_btn.bind("<Leave>", lambda e: save_add_btn.config(bg="#2196F3"))

    def save_and_add_to_talent_pool(self, entries, parent_window):
        reason_window = tk.Toplevel(self.root)
        reason_window.title("加入理由")
        reason_window.geometry("400x300")
        reason_window.configure(bg="#F0F0F0")
        self.center_window(reason_window)
        reason_window.transient(self.root)
        reason_window.grab_set()

        tk.Label(reason_window, text="请输入加入理由：", font=("Roboto", 10), bg="#F0F0F0").pack(pady=5)
        reason_text = scrolledtext.ScrolledText(reason_window, height=10, width=40, wrap=tk.WORD, font=("Roboto", 10), bd=1, relief="solid", highlightbackground="#CCCCCC", highlightthickness=1)
        reason_text.pack(pady=5)

        button_frame = tk.Frame(reason_window, bg="#F0F0F0")
        button_frame.pack(pady=10)
        confirm_btn = tk.Button(button_frame, text="确认加入", command=lambda: self.confirm_save_and_add(entries, reason_text.get("1.0", tk.END).strip(), reason_window, parent_window), font=("Roboto", 10), bg="#2196F3", fg="white", bd=0, relief="flat", padx=10, pady=5)
        confirm_btn.pack(side=tk.LEFT, padx=5)
        confirm_btn.bind("<Enter>", lambda e: confirm_btn.config(bg="#1976D2"))
        confirm_btn.bind("<Leave>", lambda e: confirm_btn.config(bg="#2196F3"))
        cancel_btn = tk.Button(button_frame, text="取消", command=reason_window.destroy, font=("Roboto", 10), bg="#FF9800", fg="white", bd=0, relief="flat", padx=10, pady=5)
        cancel_btn.pack(side=tk.LEFT, padx=5)
        cancel_btn.bind("<Enter>", lambda e: cancel_btn.config(bg="#F57C00"))
        cancel_btn.bind("<Leave>", lambda e: cancel_btn.config(bg="#FF9800"))

    def confirm_save_and_add(self, entries, reason, reason_window, parent_window):
        data = [
            entries["真实姓名"].get() or None,
            entries["性别"].get() or None,
            entries["年龄"].get() or None,
            entries["身份证号"].get() or None,
            entries["手机号"].get() or None,
            entries["省份"].get() or None,
            entries["城市"].get() or None,
            None,
            entries["昵称"].get() or None,
            entries["学历"].get() or None,
            entries["政治面貌"].get() or None,
            entries["个人职业"].get() or None,
            entries["分会职务"].get() or None,
            entries["在职状态"].get() or None,
            entries["加入组织时间"].get() or None,
            entries["跟捐天数"].get() or None,
            entries["家庭住址"].get() or None,
            entries["个人简历"].get("1.0", tk.END).strip() or None,
            self.photo_path.get() or None,
        ]
        if not data[0] or not data[4]:
            messagebox.showerror("错误", "真实姓名和手机号为必填项！")
            return
        try:
            person_id, message, error = save_and_add_to_talent_pool(data, reason)
            if error:
                messagebox.showerror("错误", error)
            else:
                self.refresh_data()
                if self.talent_window and self.talent_window.winfo_exists():
                    self.refresh_talent_list()
                messagebox.showinfo("提示", message)
                reason_window.destroy()
                parent_window.destroy()
        except Exception as e:
            messagebox.showerror("错误", f"保存失败：{str(e)}")

    def export_person_data(self, person, from_talent):
        default_filename = f"{person[1]}_详细信息.pdf"
        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=default_filename
        )
        if not file_path:
            messagebox.showinfo("提示", "取消导出")
            return
        message, error = export_person_data(person, from_talent, file_path)
        if error:
            messagebox.showerror("错误", error)
        else:
            messagebox.showinfo("成功", message)

    def close_add_person_window(self):
        if self.add_person_window:
            self.add_person_window.destroy()
            self.add_person_window = None

    def close_edit_person_window(self):
        if self.edit_person_window:
            self.edit_person_window.destroy()
            self.edit_person_window = None

if __name__ == "__main__":
    root = tk.Tk()
    app = HRManagementApp(root)
    root.mainloop()